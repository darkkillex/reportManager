import re
import datetime
import openpyxl
import pandas as pd
from openpyxl import Workbook, styles
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import constants
import utilities



list_of_sum = []

pd.set_option('display.max_columns', None)


def create_df_data_struct_report_pdl_check(df):
    df = df[["Macro area", "Tipologia attività", "Esito check"]]
    return df


def create_df_data_struct_report_pdl(df):
    df = df[["Macro Area", "Tipologia Attività"]]
    return df


def replace_comma_specific_part(string):
    if "CND (spessom,liquidi pen.,tecnografie,ultrasuoni)" in string:
        # Trova la parte specifica tra parentesi
        specific_part = re.search(r'CND \(spessom,liquidi pen.,tecnografie,ultrasuoni\)', string).group(0)
        # Sostituisci le virgole con la barra "/" solo nella parte specifica
        specific_part_replaced = specific_part.replace(',', '/')
        # Sostituisci la parte specifica nella stringa originale
        string = string.replace(specific_part, specific_part_replaced)
    return string


def find_priority_type(types, priority_list):
    #types = filter(lambda item: item != '', types)
    types = types.split(",")  # Split the types separated by commas into a list
    priority_type = None

    for priority in priority_list:
        for type in types:
            if type.strip() == priority:
                priority_type = priority
                break
        if priority_type:
            break

    return priority_type


def format_excel_sheet(ws):
   # Imposta la prima riga in grassetto (bloccata) e colorata con la palette
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=1), start=1):
        for col_index, cell in enumerate(row, start=1):
            cell.font = Font(bold=True)
            if row_index <= len(constants.COLOR_PALETTE):  # Ensure we have a color in the palette
                cell.fill = PatternFill(start_color=constants.COLOR_PALETTE[row_index - 1], end_color=constants.COLOR_PALETTE[row_index - 1], fill_type='solid')

        ws.freeze_panes = ws.cell(row=2, column=1)  # Blocca la prima riga

    last_row_index = ws.max_row
    for cell in ws[last_row_index]:
        cell.font = Font(bold=True)
    # Imposta i filtri per tutte le colonne
    ws.auto_filter.ref = ws.dimensions

    # Imposta la larghezza delle colonne in base alla lunghezza massima dei valori
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 5  # Add some padding for better visibility

    # Colora ogni riga successiva con un colore diverso dalla palette
    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        color_index = row_index % len(constants.COLOR_PALETTE)  # Use modulo to repeat colors
        for cell in row:
            cell.fill = PatternFill(start_color=constants.COLOR_PALETTE[color_index], end_color=constants.COLOR_PALETTE[color_index], fill_type='solid')



def create_check_sheets(prioritized_types, df_original, wb):
    df_original = df_original.sort_values(by=['Macro area'], ascending=True)
    # Creare fogli Excel separati per ciascuna tipologia di "Esito"
    for esito in df_original['Esito check'].unique():
        ws = wb.create_sheet(title='Check_PDL-' + esito)
        df_excel = pd.DataFrame()

        # Filtrare il DataFrame originale per la tipologia di "Esito"
        df_filtered = df_original[df_original['Esito check'] == esito]

        # Inserire le tipologie di attività nella prima colonna
        df_excel['Tipologia attività'] = prioritized_types

        # Iterare attraverso le "macro aree" univoche
        all_macro_areas = df_original['Macro area'].unique()

        # Inserisci tutte le macro aree come colonne, inizializzate a 0
        for macro_area in all_macro_areas:
            df_excel[macro_area] = 0

        # Eseguire il conteggio delle occorrenze
        counts_df = df_filtered.groupby(['Macro area', 'Tipologia attività']).size().reset_index(name='Counts')

        # Aggiungere i conteggi al DataFrame Excel
        for index, row in counts_df.iterrows():
            macro_area = row['Macro area']
            activity_type = row['Tipologia attività']
            count = row['Counts']

            # Verifica se c'è una corrispondenza tra "macro area" e "tipologia attività"
            if macro_area in df_excel.columns and activity_type in df_excel['Tipologia attività'].values:
                df_excel.loc[df_excel['Tipologia attività'] == activity_type, macro_area] = count

        df_excel = df_excel.replace('', 0)
        # Calcola e aggiungi la riga con la somma delle colonne (tralasciando la prima colonna)
        sums = df_excel.iloc[:, 1:].sum()
        df_sums = pd.DataFrame([['TOTALE'] + sums.tolist()], columns=df_excel.columns)
        list_of_sum.append(df_sums.iloc[0, 1:].sum())
        df_excel = pd.concat([df_excel, df_sums], ignore_index=True)
        for r in dataframe_to_rows(df_excel, index=False, header=True):
            ws.append(r)
        format_excel_sheet(ws)


def create_pdl_sheets(prioritized_types, df_original_pdl, wb, sheet_label):
    df_original_pdl = df_original_pdl.sort_values(by=['Macro Area'], ascending=True)
    #df_original_pdl = df_original_pdl.dropna(subset=['Tipologia Attività'])
    ws = wb.create_sheet(title=sheet_label)
    df_excel = pd.DataFrame()
    # Inserire le tipologie di attività nella prima colonna
    df_excel['Tipologia Attività'] = prioritized_types

    # Iterare attraverso le "macro aree" univoche
    all_macro_areas = df_original_pdl['Macro Area'].unique()

    # Inserisci tutte le macro aree come colonne, inizializzate a 0
    for macro_area in all_macro_areas:
        df_excel[macro_area] = 0

    # Eseguire il conteggio delle occorrenze
    counts_df = df_original_pdl.groupby(['Macro Area', 'Tipologia Attività']).size().reset_index(name='Counts_pdl')

    # Aggiungere i conteggi al DataFrame Excel
    for index, row in counts_df.iterrows():
        macro_area = row['Macro Area']
        activity_type = row['Tipologia Attività']
        count = row['Counts_pdl']

        # Verifica se c'è una corrispondenza tra "macro area" e "tipologia attività"
        if macro_area in df_excel.columns and activity_type in df_excel['Tipologia Attività'].values:
            df_excel.loc[df_excel['Tipologia Attività'] == activity_type, macro_area] = count

    df_excel = df_excel.replace('', 0)
    # Calcola e aggiungi la riga con la somma delle colonne (tralasciando la prima colonna)
    sums = df_excel.iloc[:, 1:].sum()
    df_sums = pd.DataFrame([['TOTALE'] + sums.tolist()], columns=df_excel.columns)
    list_of_sum.append(df_sums.iloc[0, 1:].sum())
    df_excel = pd.concat([df_excel, df_sums], ignore_index=True)
    for r in dataframe_to_rows(df_excel, index=False, header=True):
        ws.append(r)
    format_excel_sheet(ws)


def define_report_summary_part(wb):
    # Create a new sheet
    riepilogo_sheet = wb.create_sheet("Riepilogo")
    # Calculate the sum of all data in the "TOTALE" row in the PDL sheet

    # Write the sum to the cell in the "Riepilogo" sheet

    riepilogo_sheet['A1'] = "Check PDL Positivi"
    riepilogo_sheet['A2'] = "Check PDL con Problematiche"
    riepilogo_sheet['A3'] = "Check PDL Azione Preventiva"
    riepilogo_sheet['A4'] = "Check PDL Stop Work"
    riepilogo_sheet['A5'] = "TOTALE CHECK PDL"
    riepilogo_sheet['B1'] = list_of_sum[0]
    riepilogo_sheet['B2'] = list_of_sum[1]
    riepilogo_sheet['B3'] = list_of_sum[2]
    riepilogo_sheet['B4'] = list_of_sum[3]
    riepilogo_sheet['B5'] = sum(list_of_sum[:4])

    riepilogo_sheet['A6'] = "N° PDL Protocollati"
    riepilogo_sheet['A7'] = "N° PDL Autorizzati"
    riepilogo_sheet['A8'] = "Incidenza"
    riepilogo_sheet['B6'] = list_of_sum[4]
    riepilogo_sheet['B7'] = list_of_sum[5]
    riepilogo_sheet['B8'] = '{:.2%}'.format(list_of_sum[5] / list_of_sum[4])


def define_report_check_part(source_file, column_name, wb):
    df_pdl_check = utilities.load_xlsx_file(source_file)
    df_final_pdl_check = utilities.clean_df(df_pdl_check)
    df_final_pdl_check = df_final_pdl_check.dropna(subset=['Tipologia attività'])
    df_temp_pdl_check = create_df_data_struct_report_pdl_check(df_final_pdl_check)
    df_ultimate_pdl_check = df_temp_pdl_check.copy()
    df_ultimate_pdl_check[column_name] = df_ultimate_pdl_check[column_name].\
        apply(replace_comma_specific_part)
    df_ultimate_pdl_check[column_name] = df_ultimate_pdl_check[column_name].\
        apply(lambda x: find_priority_type(x, priority_list=constants.LIST_PRIORITY_PDL_AND_CHECK))
    create_check_sheets(constants.LIST_PRIORITY_PDL_AND_CHECK, df_ultimate_pdl_check, wb)


def define_report_pdl_part(source_file, column_name, wb, sheet_label):
    df_pdl = utilities.load_xlsx_file(source_file)
    df_final_pdl = utilities.clean_df(df_pdl)
    df_final_pdl = df_final_pdl.dropna(subset=['Tipologia Attività'])
    df_temp_pdl = create_df_data_struct_report_pdl(df_final_pdl)
    df_ultimate_pdl = df_temp_pdl.copy()
    df_ultimate_pdl[column_name] = df_ultimate_pdl[column_name].\
        apply(replace_comma_specific_part)
    df_ultimate_pdl[column_name] = df_ultimate_pdl[column_name].\
        apply(lambda x: find_priority_type(x, priority_list=constants.LIST_PRIORITY_PDL_AND_CHECK))
    create_pdl_sheets(constants.LIST_PRIORITY_PDL_AND_CHECK, df_ultimate_pdl, wb, sheet_label)


def create_excel_output(output_file):
    wb = Workbook()
    #Outcome of Check PDL
    define_report_check_part(constants.PDL_CHECK, 'Tipologia attività', wb)
    #Outcome of PDL Prot
    define_report_pdl_part(constants.PDL_PROT, 'Tipologia Attività', wb, "PDL-Protocollati")
    #Outcome of PDL Aut
    define_report_pdl_part(constants.PDL_AUT, 'Tipologia Attività', wb, "PDL-Autorizzati")
    define_report_summary_part(wb)
    # Rimuovi il foglio di lavoro predefinito
    wb.remove(wb.active)
    # Salva il Workbook completo in un file Excel
    wb.save(output_file)


def generate_report_label(df):
    # Extract the first column containing the dates as strings
    if not df.empty:
        today = datetime.date.today()
        week = today.strftime("%U")  # Numero della settimana
        year = today.strftime("%Y")  # Anno corrente
        report_label = "assets/report_pdl_check_pdl/" + constants.LABEL_REPORT_PDL_CHECK + week + "-" + year + ".xlsx"

    return report_label


def run_scripts_report_pdl_check():
    create_excel_output("assets/report_pdl_check_pdl/Report PDL-Check PDL-Sett.XX-XXXX.xlsx")
    print("Report PDL/Check PDL has been generated in the /assets/report_pdl_check_pdl")
