from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

import constants
import utilities


wb_adr_in_out, wb_not_adr_in_out, wb_adr_vob_pob, wb_not_adr_vob_pob = Workbook(), Workbook(), Workbook(), Workbook()
adr_sheet_in_out = {}
not_adr_sheet_in_out = {}
adr_sheet_vob_pob = {}
not_adr_sheet_vob_pob = {}


def write_on_xlsx_sheet_file(wb, sheets):
    for group_name, group_data in sheets.items():
        sheet = wb.create_sheet(title=group_name)
        # adjust the width of the columns
        for col in range(1, len(group_data.columns) + 1):
            sheet.column_dimensions[sheet.cell(1, col).column_letter].width = 40
        # writa the data
        for row in dataframe_to_rows(group_data, index=False, header=True):
            sheet.append(row)
        # Set to Bold the first line
        for cell in sheet['2']:
            cell.font = Font(bold=True)


def create_report(wb, sheet, label):
    write_on_xlsx_sheet_file(wb, sheet)
    wb.remove(wb['Sheet'])
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Elimina la prima riga (Header)
        sheet.delete_rows(1)
        # Imposta i filtri per tutte le colonne
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'
    wb.save(label)


def generate_report_label(df, label):
    # Extract the first column containing the dates as strings
    if not df.empty:
        cell_value = df.iloc[1, 1] # get the date from
        date_parts = cell_value.split(' ')
        date_string = date_parts[0]
        # Replace the symbol "/" with "-"
        date_string = date_string.replace('/', '-') # now we have the date
        report_label = "assets/in_out_vob/" + label + date_string + ".xlsx"

    return report_label


def create_df_data_struct_report_in_out_pob(df, type_of_report):
    if type_of_report == "IN_OUT":
        #switch data between col1 and col2
        temp_IN_OUT = df['Momento'].copy()  # Copia i dati di Colonna1 in una variabile temporanea
        df['Momento'] = df['Appaltatore']  # Sovrascrive i dati di Colonna1 con quelli di Colonna2
        df['Appaltatore'] = temp_IN_OUT  # Sovrascrive i dati di Colonna2 con quelli dalla variabile temporanea
        df = df.rename(columns={'Momento': 'Azienda Appaltatrice'})
        df = df.rename(columns={'Appaltatore': 'Momento'})
    if type_of_report == "VOB_POB":
        df = df[constants.LIST_OF_LABELS_VOB_POB]
        df = df.rename(columns={'Appaltatore': 'Azienda Appaltatrice'})


    # cuts the number of characters in the excel sheet label (max 31 char supported)
    df['Azienda Appaltatrice'] = df['Azienda Appaltatrice'].str.slice(0, 31)
    return df


def populate_sheets(df, adr_sheet, not_adr_sheet):
    for group, group_data in df.groupby('Azienda Appaltatrice'):
        temp_adr = group_data[group_data['Tipologia'].str.contains('ADR', case=False, na=False)]
        temp_not_adr = group_data[~group_data['Tipologia'].str.contains('ADR', case=False, na=False)]
        if not temp_adr.empty:
             adr_sheet[group] = temp_adr
        if not temp_not_adr.empty:
             not_adr_sheet[group] = temp_not_adr

def run_scripts_report_in_out_pob():
    # let's go work on DFs
    df_in_out = utilities.load_xlsx_file(constants.IN_OUT)
    df_vob_pob = utilities.load_xlsx_file(constants.VOB_POB)

    if df_in_out is not None:
        df_final_in_out = utilities.clean_df(df_in_out)
        df_ultimate_in_out = create_df_data_struct_report_in_out_pob(df_final_in_out, type_of_report=constants.REPORT_IN_OUT)
        populate_sheets(df_ultimate_in_out, adr_sheet_in_out, not_adr_sheet_in_out)
        create_report(wb_adr_in_out, adr_sheet_in_out,
                      generate_report_label(df_ultimate_in_out, constants.LABEL_REPORT_IN_OUT_ADR))
        create_report(wb_not_adr_in_out, not_adr_sheet_in_out, generate_report_label(df_ultimate_in_out,
                                                                                     constants.LABEL_REPORT_IN_OUT_NOT_ADR))
        print("IN_OUT report has been generated in the /assets/in_out_vob folder!")
    if df_vob_pob is not None:
        df_final_vob_pob = utilities.clean_df(df_vob_pob)
        df_ultimate_vob_pob = create_df_data_struct_report_in_out_pob(df_final_vob_pob, type_of_report=constants.REPORT_VOB_POB)
        populate_sheets(df_ultimate_vob_pob, adr_sheet_vob_pob, not_adr_sheet_vob_pob)
        #print(df_ultimate_vob_pob.head())
        create_report(wb_adr_vob_pob, adr_sheet_vob_pob, generate_report_label(df_ultimate_in_out, #in this case we get the df_in_out to get the right date of extraction
                                                                               constants.LABEL_REPORT_VOB_POB_ADR))
        create_report(wb_not_adr_vob_pob, not_adr_sheet_vob_pob, generate_report_label(df_ultimate_in_out, #in this case we get the df_in_out to get the right date of extraction
                                                                                       constants.LABEL_REPORT_VOB_POB_NOT_ADR))
        print("VOB report has been generated in the /assets/in_out_vob folder!")


